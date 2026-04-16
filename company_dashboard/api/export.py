"""One-shot MIS exports — Excel workbook and A4 PDF report.

Excel (.xlsx) delivers one sheet per dashboard page:
  Overview KPIs · Monthly Revenue · Revenue by Category · P&L Waterfall ·
  SKU Assortment · Customer Segments.

PDF is an A4 report (portrait) laid out for printing / board packs: cover
page, CM ladder summary, P&L line-by-line, monthly revenue table, top-10
categories, top-50 SKUs, and per-segment snapshot.

Both share the same data pull — they just differ in presentation. Each
section is self-describing (titles, period, tax-mode, allocation basis) so
an export audits cleanly standalone without opening the dashboard.
"""

from __future__ import annotations

import base64
import html as _html
import io
import os
from datetime import datetime, timezone

import frappe
from frappe import _
from frappe.utils.pdf import get_pdf
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from company_dashboard.api.mis import (
	_require_mis_access,
	get_customer_segments,
	get_overview,
	get_sku_assortment,
)

# Green header band matching the dashboard's brand — WCAG-AA against white text.
_HEADER_FILL = PatternFill("solid", fgColor="1F5F33")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SECTION_FONT = Font(bold=True, color="1A1810", size=12)
_MUTED_FONT = Font(color="7A6F5D", size=10, italic=True)
_NUM_FMT_INR = '"₹"#,##0'
_NUM_FMT_PCT = "0.00%"
_NUM_FMT_INT = "#,##0"


@frappe.whitelist(allow_guest=False)
def download_mis_workbook(
	tax_mode: str = "incl",
	from_date: str | None = None,
	to_date: str | None = None,
	intent: str = "all",
) -> None:
	"""Whitelisted endpoint — streams the workbook back to the browser as a download.

	Frappe's ``frappe.response`` hook handles the file mechanics: we set ``filename``,
	``filecontent``, and ``type="binary"`` and Frappe returns the right Content-Type +
	Content-Disposition headers automatically.
	"""
	_require_mis_access()

	# Gather all four pages' data with the caller's filter settings. Each call does
	# its own access check, tax-mode validation, and period resolution.
	overview = get_overview(tax_mode=tax_mode, from_date=from_date, to_date=to_date)
	assortment = get_sku_assortment(
		tax_mode=tax_mode, from_date=from_date, to_date=to_date, intent=intent
	)
	segments = get_customer_segments(
		tax_mode=tax_mode, from_date=from_date, to_date=to_date, intent=intent
	)

	wb = Workbook()
	# Drop the default blank sheet.
	default = wb.active
	wb.remove(default)

	_write_overview_sheet(wb, overview)
	_write_monthly_sheet(wb, overview)
	_write_categories_sheet(wb, overview)
	_write_pnl_sheet(wb, overview)
	_write_assortment_sheet(wb, assortment)
	_write_segments_sheet(wb, segments)

	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)

	period_label = overview["period"]["label"].replace(" ", "_")
	filename = f"MIS_{overview['company'].replace(' ', '_')}_{period_label}.xlsx"
	frappe.response.update(
		{
			"filename": filename,
			"filecontent": buf.getvalue(),
			"type": "binary",
			"display_content_as": "attachment",
		}
	)


# ── Sheet writers ─────────────────────────────────────────────────────────────


def _write_overview_sheet(wb: Workbook, res: dict) -> None:
	ws = wb.create_sheet("Overview KPIs")
	kpi = res["kpi"]
	ws.append(["MIS Overview", res["company"]])
	ws["A1"].font = _SECTION_FONT
	ws.append([res["period"]["label"], res["tax_mode"].upper()])
	ws["A2"].font = _MUTED_FONT
	ws.append([])

	_section(ws, "Revenue")
	_kv_row(ws, "Total Revenue", kpi["total_revenue"], _NUM_FMT_INR)
	_kv_row(ws, "Total Revenue (prior)", kpi["total_revenue_prior"], _NUM_FMT_INR)
	_kv_row(ws, "B2C Revenue", kpi["b2c_revenue"], _NUM_FMT_INR)
	_kv_row(ws, "B2B Revenue", kpi["b2b_revenue"], _NUM_FMT_INR)
	_kv_row(ws, "Unclassified Revenue", kpi["unclassified_revenue"], _NUM_FMT_INR)
	ws.append([])

	_section(ws, "Gross Margin")
	_kv_row(ws, "Gross Profit", kpi["gross_profit"], _NUM_FMT_INR)
	_kv_row(ws, "Gross Margin %", kpi["gross_margin_pct"] / 100, _NUM_FMT_PCT)
	_kv_row(ws, "Valuation Coverage %", kpi["gross_margin_coverage_pct"] / 100, _NUM_FMT_PCT)
	_kv_row(ws, "B2C GM %", kpi["b2c_gross_margin_pct"] / 100, _NUM_FMT_PCT)
	_kv_row(ws, "B2B GM %", kpi["b2b_gross_margin_pct"] / 100, _NUM_FMT_PCT)
	ws.append([])

	_section(ws, "Contribution Margin ladder")
	for label, abs_val, pct_val in [
		("CM1 (Gross Margin)", kpi["cm1"], kpi["cm1_pct"]),
		("CM2 (before marketing)", kpi["cm2"], kpi["cm2_pct"]),
		("CM3 (after marketing)", kpi["cm3"], kpi["cm3_pct"]),
		("EBITDA", kpi["ebitda"], kpi["ebitda_pct"]),
		("EBIT", kpi["ebit"], kpi["ebit_pct"]),
		("PBT", kpi["pbt"], kpi["pbt_pct"]),
		("PAT", kpi["pat"], kpi["pat_pct"]),
	]:
		row = ws.max_row + 1
		ws.cell(row=row, column=1, value=label)
		ws.cell(row=row, column=2, value=abs_val).number_format = _NUM_FMT_INR
		ws.cell(row=row, column=3, value=pct_val / 100).number_format = _NUM_FMT_PCT
	ws.append([])

	_section(ws, "Activity")
	_kv_row(ws, "Active SKUs", kpi["active_skus"], _NUM_FMT_INT)
	_kv_row(ws, "Average MRP/SKU", kpi["avg_mrp"], _NUM_FMT_INR)
	ws.append([])

	_write_provenance(ws, res, "CM1 = Gross Margin · CM2 = −freight/pkg/txn/schemes · CM3 = CM2 − marketing")
	_autofit(ws, [36, 18, 18])


def _write_monthly_sheet(wb: Workbook, res: dict) -> None:
	ws = wb.create_sheet("Monthly Revenue")
	rows = res.get("monthly", [])
	_header_row(ws, ["Month", "Year", "B2C (₹L)", "B2B (₹L)", "Other (₹L)", "Total (₹L)"])
	for r in rows:
		ws.append([r["month"], r["year"], r["b2c"], r["b2b"], r["other"], r["total"]])
	# Totals
	last = ws.max_row
	if rows:
		total_row = last + 1
		ws.cell(row=total_row, column=1, value="Total").font = _SECTION_FONT
		for col in (3, 4, 5, 6):
			letter = get_column_letter(col)
			ws.cell(
				row=total_row, column=col, value=f"=SUM({letter}2:{letter}{last})"
			).font = _SECTION_FONT
	_write_provenance(ws, res, "Revenue in ₹ lakhs. 'Other' = items without a B2C/B2B custom_sales_intent.")
	_autofit(ws, [10, 8, 12, 12, 12, 12])


def _write_categories_sheet(wb: Workbook, res: dict) -> None:
	ws = wb.create_sheet("Revenue by Category")
	rows = res.get("categories", [])
	_header_row(ws, ["Category", "SKU Count", "Revenue (₹L)", "% of Shown"])
	total = sum(r["revenue"] for r in rows) or 1
	for r in rows:
		share = r["revenue"] / total
		ws.append([r["category"], r["sku_count"], r["revenue"], share])
		ws.cell(row=ws.max_row, column=4).number_format = _NUM_FMT_PCT
	total_row = ws.max_row + 1
	ws.cell(row=total_row, column=1, value="Total").font = _SECTION_FONT
	ws.cell(row=total_row, column=2, value=sum(r["sku_count"] for r in rows)).font = _SECTION_FONT
	ws.cell(row=total_row, column=3, value=total).font = _SECTION_FONT
	ws.cell(row=total_row, column=4, value=1).number_format = _NUM_FMT_PCT
	_write_provenance(ws, res, "Top 10 Item Groups by revenue; rest collapsed into 'Other'.")
	_autofit(ws, [28, 12, 14, 12])


def _write_pnl_sheet(wb: Workbook, res: dict) -> None:
	ws = wb.create_sheet("P&L Waterfall")
	kpi = res["kpi"]
	c2 = res["cm2_inputs"]
	c3 = res["cm3_inputs"]
	li = res["ladder_inputs"]
	rev = kpi["total_revenue"] or 1

	_header_row(ws, ["Line Item", "Amount (₹)", "% of Revenue", "Detail"])

	def line(label: str, amount: float, is_tier: bool = False, detail: str = "") -> None:
		row = ws.max_row + 1
		ws.cell(row=row, column=1, value=label)
		ws.cell(row=row, column=2, value=amount).number_format = _NUM_FMT_INR
		ws.cell(row=row, column=3, value=amount / rev).number_format = _NUM_FMT_PCT
		ws.cell(row=row, column=4, value=detail)
		if is_tier:
			for col in range(1, 5):
				ws.cell(row=row, column=col).font = Font(bold=True)

	line("Gross Revenue", kpi["total_revenue"], is_tier=True)
	line("    COGS (incoming_rate × qty)", -(kpi["total_revenue"] - kpi["gross_profit"]), detail="includes primary packaging via BOM")
	line("CM1 · Gross Margin", kpi["cm1"], is_tier=True)

	line("    Outbound Freight", -c2["freight_total"], detail=f"{len(c2['freight_accounts'])} accounts")
	line("    Packaging Overhead", -c2["packaging_total"], detail=f"{len(c2['packaging_accounts'])} accounts")
	line("    Transaction Fees & Commissions", -c2["commission_total"], detail=f"{len(c2['commission_accounts'])} accounts")
	line("    Trade Schemes", -c2["scheme_total"], detail=f"{len(c2['scheme_accounts'])} accounts")
	line("CM2 · Before marketing", kpi["cm2"], is_tier=True)

	line("    Marketing", -c3["marketing_total"], detail=f"{len(c3['marketing_accounts'])} accounts")
	line("CM3 · After marketing", kpi["cm3"], is_tier=True)

	line("    Salary & Payroll", -li["salary_total"], detail=f"{len(li['salary_accounts'])} accounts")
	line("    Rent & Lease", -li["rent_total"], detail=f"{len(li['rent_accounts'])} accounts")
	line("    Utilities & Communications", -li["utility_total"], detail=f"{len(li['utility_accounts'])} accounts")
	line("    Admin & General Overhead", -li["admin_total"], detail=f"{len(li['admin_accounts'])} accounts")
	line("EBITDA", kpi["ebitda"], is_tier=True)

	line("    Depreciation & Amortization", -li["da_total"], detail=f"{len(li['da_accounts'])} accounts")
	line("EBIT", kpi["ebit"], is_tier=True)

	line("    Interest & Finance Cost", -li["interest_total"], detail=f"{len(li['interest_accounts'])} accounts")
	line("PBT", kpi["pbt"], is_tier=True)

	line("    Income Tax Expense", -li["tax_total"], detail=f"{len(li['tax_accounts'])} accounts")
	line("PAT", kpi["pat"], is_tier=True)

	ws.append([])
	_write_provenance(ws, res, "Line-by-line P&L from Sales Invoice + GL. All CM tiers computed tax-exclusive.")
	_autofit(ws, [38, 16, 14, 40])


def _write_assortment_sheet(wb: Workbook, res: dict) -> None:
	ws = wb.create_sheet("SKU Assortment")
	rows = res.get("rows", [])
	month_keys = res.get("month_keys", [])
	base_cols = [
		"#", "SKU", "Item Name", "Item Group", "Intent",
		"Units", "Revenue (₹)", "COGS (₹)", "GM %",
		"Freight (₹)", "Packaging (₹)", "Txn Fees (₹)", "Scheme (₹)", "Marketing (₹)",
		"CM2 %", "CM3 %",
	]
	_header_row(ws, base_cols + month_keys)

	for i, r in enumerate(rows, start=1):
		row = [
			i, r["item_code"], r["item_name"], r["item_group"], r["intent"],
			r["qty"], r["revenue"], r["cogs"], r["gm_pct"] / 100,
			r["alloc_freight"], r["alloc_packaging"], r["alloc_commission"],
			r["alloc_scheme"], r["alloc_marketing"],
			r["cm2_pct"] / 100, r["cm3_pct"] / 100,
		]
		for mk in month_keys:
			row.append(r["monthly"].get(mk, 0.0))
		ws.append(row)
		rnum = ws.max_row
		# Number formats
		for col in (6,):
			ws.cell(row=rnum, column=col).number_format = _NUM_FMT_INT
		for col in (7, 8, 10, 11, 12, 13, 14):
			ws.cell(row=rnum, column=col).number_format = _NUM_FMT_INR
		for col in (9, 15, 16):
			ws.cell(row=rnum, column=col).number_format = _NUM_FMT_PCT
		for col_idx in range(len(base_cols) + 1, len(base_cols) + 1 + len(month_keys)):
			ws.cell(row=rnum, column=col_idx).number_format = _NUM_FMT_INR

	ws.freeze_panes = "F2"  # freeze first 5 cols + header
	_write_provenance(
		ws,
		res,
		f"Allocations: freight {res['allocation']['freight_pct']:.2f}% · "
		f"pkg {res['allocation']['packaging_pct']:.2f}% · "
		f"txn {res['allocation']['commission_pct']:.2f}% · "
		f"marketing {res['allocation']['marketing_pct']:.2f}% of revenue.",
	)
	_autofit(ws, [5, 14, 32, 18, 14, 10, 14, 14, 9, 13, 13, 13, 12, 14, 9, 9] + [11] * len(month_keys))


def _write_segments_sheet(wb: Workbook, res: dict) -> None:
	ws = wb.create_sheet("Customer Segments")
	rows = res.get("segments", [])
	_header_row(
		ws,
		[
			"Segment", "Customers", "Invoices", "Revenue (₹)",
			"% Mix", "YoY %", "AOV (₹)", "GP (₹)", "GM %",
		],
	)
	total_rev = res["totals"]["revenue"] or 1
	for r in rows:
		share = r["revenue"] / total_rev
		yoy = (r["yoy_pct"] / 100) if r["yoy_pct"] is not None else None
		ws.append([
			r["segment"], r["customers"], r["invoices"], r["revenue"],
			share, yoy, r["aov"], r["gp"], r["gm_pct"] / 100,
		])
		rnum = ws.max_row
		ws.cell(row=rnum, column=2).number_format = _NUM_FMT_INT
		ws.cell(row=rnum, column=3).number_format = _NUM_FMT_INT
		ws.cell(row=rnum, column=4).number_format = _NUM_FMT_INR
		ws.cell(row=rnum, column=5).number_format = _NUM_FMT_PCT
		ws.cell(row=rnum, column=6).number_format = _NUM_FMT_PCT
		ws.cell(row=rnum, column=7).number_format = _NUM_FMT_INR
		ws.cell(row=rnum, column=8).number_format = _NUM_FMT_INR
		ws.cell(row=rnum, column=9).number_format = _NUM_FMT_PCT

	# Top-3 customers per segment inline, as a secondary block below.
	ws.append([])
	ws.append(["— Top customers per segment —"])
	ws.cell(row=ws.max_row, column=1).font = _SECTION_FONT
	_header_row(ws, ["Segment", "Customer", "Invoices", "Revenue (₹)", "% of Segment"])
	for r in rows:
		if not r["top_customers"]:
			continue
		seg_rev = r["revenue"] or 1
		for c in r["top_customers"][:5]:
			share = c["revenue"] / seg_rev
			ws.append([r["segment"], c["customer_name"], c["invoices"], c["revenue"], share])
			rnum = ws.max_row
			ws.cell(row=rnum, column=3).number_format = _NUM_FMT_INT
			ws.cell(row=rnum, column=4).number_format = _NUM_FMT_INR
			ws.cell(row=rnum, column=5).number_format = _NUM_FMT_PCT

	scope_note = (
		f"Scoped to {', '.join(res['scope_roots'])} (descendants rolled up)"
		if res.get("scope_roots")
		else "All Customer Groups with sales in the period"
	)
	_write_provenance(ws, res, scope_note)
	_autofit(ws, [30, 12, 12, 16, 10, 10, 14, 14, 10])


# ── Small helpers ─────────────────────────────────────────────────────────────


def _header_row(ws, labels: list[str]) -> None:
	ws.append(labels)
	for cell in ws[ws.max_row]:
		cell.font = _HEADER_FONT
		cell.fill = _HEADER_FILL
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _section(ws, label: str) -> None:
	ws.append([label])
	ws.cell(row=ws.max_row, column=1).font = _SECTION_FONT


def _kv_row(ws, key: str, value: float | int | str, fmt: str | None = None) -> None:
	ws.append([key, value])
	if fmt:
		ws.cell(row=ws.max_row, column=2).number_format = fmt


def _write_provenance(ws, res: dict, note: str) -> None:
	ws.append([])
	ws.append(["Provenance"])
	ws.cell(row=ws.max_row, column=1).font = _SECTION_FONT
	ws.append(["Company", res.get("company", "—")])
	ws.append(["Period", res.get("period", {}).get("label", "—")])
	ws.append(["Tax mode", res.get("tax_mode", "—")])
	ws.append(["Exported", datetime.now(timezone.utc).isoformat(timespec="seconds")])
	ws.append(["Notes", note])
	ws.cell(row=ws.max_row, column=2).alignment = Alignment(wrap_text=True)
	for col in (1,):
		for row in range(ws.max_row - 5, ws.max_row + 1):
			ws.cell(row=row, column=col).font = _MUTED_FONT


def _autofit(ws, widths: list[int]) -> None:
	for i, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════════════════════
# PDF REPORT — A4 portrait, green+beige theme matching the dashboard.
# ══════════════════════════════════════════════════════════════════════════════


@frappe.whitelist(allow_guest=False)
def download_mis_pdf(
	tax_mode: str = "incl",
	from_date: str | None = None,
	to_date: str | None = None,
	intent: str = "all",
) -> None:
	"""Render an A4 MIS report as PDF. Same data the dashboard uses — board-ready layout."""
	_require_mis_access()

	overview = get_overview(tax_mode=tax_mode, from_date=from_date, to_date=to_date)
	assortment = get_sku_assortment(
		tax_mode=tax_mode, from_date=from_date, to_date=to_date, intent=intent
	)
	segments = get_customer_segments(
		tax_mode=tax_mode, from_date=from_date, to_date=to_date, intent=intent
	)

	html_doc = _build_pdf_html(overview, assortment, segments, intent=intent)
	pdf_bytes = get_pdf(
		html_doc,
		{
			"page-size": "A4",
			"margin-top": "14mm",
			"margin-bottom": "16mm",
			"margin-left": "14mm",
			"margin-right": "14mm",
			"encoding": "UTF-8",
			"print-media-type": None,
			"disable-smart-shrinking": None,
		},
	)

	period_label = overview["period"]["label"].replace(" ", "_").replace("·", "")
	filename = f"MIS_{overview['company'].replace(' ', '_')}_{period_label}.pdf"
	frappe.response.update(
		{
			"filename": filename,
			"filecontent": pdf_bytes,
			"type": "pdf",
			"display_content_as": "attachment",
		}
	)


def _build_pdf_html(overview: dict, assortment: dict, segments: dict, intent: str) -> str:
	"""Inline-CSS HTML → wkhtmltopdf. Keeping styles in-file avoids static-file path
	quirks on bench deploys; the tradeoff is verbosity we accept once per export."""
	kpi = overview["kpi"]
	c2 = overview["cm2_inputs"]
	c3 = overview["cm3_inputs"]
	li = overview["ladder_inputs"]
	rev = kpi["total_revenue"] or 1

	now = datetime.now(timezone.utc).strftime("%d %b %Y · %H:%M UTC")
	intent_label = {"all": "All items", "b2c": "B2C items only", "b2b": "B2B items only"}.get(intent, "All items")
	tax_label = "Tax-inclusive" if overview["tax_mode"] == "incl" else "Tax-exclusive"

	# ── Cover + chrome ────────────────────────────────────────────────────
	parts: list[str] = [
		"<!doctype html>",
		"<html><head><meta charset='utf-8'>",
		f"<title>MIS — {_esc(overview['company'])}</title>",
		_PDF_CSS,
		"</head><body>",
		# Cover page
		"<section class='cover'>",
		"<div class='cover-eyebrow'>Consumer Business MIS</div>",
		f"<h1 class='cover-title'>{_esc(overview['company'])}</h1>",
		f"<div class='cover-period'>{_esc(overview['period']['label'])}</div>",
		"<div class='cover-tags'>",
		f"<span class='tag tag-green'>{_esc(tax_label)}</span>",
		f"<span class='tag tag-amber'>{_esc(intent_label)}</span>",
		"</div>",
		"<div class='cover-spacer'></div>",
		"<table class='cover-kpis'><tr>",
		_cover_kpi("Revenue", _inr(kpi["total_revenue"])),
		_cover_kpi("CM1 · GM", _pct(kpi["cm1_pct"]), accent="green"),
		_cover_kpi("CM2", _pct(kpi["cm2_pct"])),
		_cover_kpi("CM3", _pct(kpi["cm3_pct"]), accent="amber" if kpi["cm3_pct"] >= 0 else "red"),
		_cover_kpi("EBITDA", _pct(kpi["ebitda_pct"]), accent="green" if kpi["ebitda_pct"] >= 0 else "red"),
		_cover_kpi("PAT", _pct(kpi["pat_pct"]), accent="green" if kpi["pat_pct"] >= 0 else "red"),
		"</tr></table>",
		f"<div class='cover-meta'>Generated {_esc(now)}</div>",
		"</section>",
	]

	# ── Page 2 · Revenue & CM ladder ──────────────────────────────────────
	parts.extend(
		[
			"<section class='page'>",
			_section_header("Revenue & Contribution Margin", overview["period"]["label"]),
			"<table class='kpi-grid'>",
			"<tr>",
			_kpi_cell("Total Revenue", _inr(kpi["total_revenue"]), _yoy_chip(kpi["total_revenue"], kpi["total_revenue_prior"])),
			_kpi_cell("B2C Revenue", _inr(kpi["b2c_revenue"]), _yoy_chip(kpi["b2c_revenue"], kpi["b2c_revenue_prior"])),
			_kpi_cell("B2B Revenue", _inr(kpi["b2b_revenue"]), _yoy_chip(kpi["b2b_revenue"], kpi["b2b_revenue_prior"])),
			"</tr>",
			"<tr>",
			_kpi_cell("Gross Profit (valuated)", _inr(kpi["gross_profit"]), f"{round(kpi['gross_margin_coverage_pct'])}% coverage"),
			_kpi_cell("Gross Margin %", _pct(kpi["gross_margin_pct"]), "tax-excl", accent="green" if kpi["gross_margin_pct"] >= 0 else "red"),
			_kpi_cell("Active SKUs", f"{int(kpi['active_skus']):,}", "items with sales", accent="blue"),
			"</tr>",
			"</table>",
			"<div class='sub-header'>Monthly revenue — B2C vs B2B</div>",
			_svg_monthly_chart(overview["monthly"]),
			"<div class='sub-header'>CM ladder — % of revenue</div>",
			_svg_cm_ladder(kpi),
			_ladder_table(kpi),
			"</section>",
		]
	)

	# ── Page 3 · P&L line-by-line ─────────────────────────────────────────
	parts.extend(
		[
			"<section class='page page-break'>",
			_section_header("P&L Waterfall", f"Revenue → PAT · {tax_label}"),
			_pnl_table(kpi, c2, c3, li, rev),
			"</section>",
		]
	)

	# ── Page 4 · Monthly + Categories ─────────────────────────────────────
	parts.extend(
		[
			"<section class='page page-break'>",
			_section_header("Monthly Revenue", "₹ Lakh"),
			_monthly_table(overview["monthly"]),
			"<div class='spacer-16'></div>",
			_section_header("Revenue by Category", "top 10 · Item Group"),
			_svg_category_bars(overview["categories"]),
			_category_table(overview["categories"]),
			"</section>",
		]
	)

	# ── Page 5+ · Top SKUs ────────────────────────────────────────────────
	parts.extend(
		[
			"<section class='page page-break'>",
			_section_header("SKU Assortment — Top 50", f"{intent_label} · sorted by revenue"),
			_sku_table(assortment["rows"][:50]),
			_allocation_footer(assortment["allocation"]),
			"</section>",
		]
	)

	# ── Page N · Customer Segments ────────────────────────────────────────
	parts.extend(
		[
			"<section class='page page-break'>",
			_section_header(
				"Customer Segments",
				f"{segments['totals']['segment_count']} segments · {intent_label}",
			),
			_svg_segment_bars(segments),
			_segment_summary_table(segments),
			_segment_top_customers(segments),
			"</section>",
		]
	)

	parts.append("</body></html>")
	return "".join(parts)


# ── Small template helpers ────────────────────────────────────────────────────


def _section_header(title: str, sub: str) -> str:
	return (
		"<div class='section-head'>"
		f"<div class='section-title'>{_esc(title)}</div>"
		f"<div class='section-sub'>{_esc(sub)}</div>"
		"</div>"
	)


def _cover_kpi(label: str, value: str, accent: str = "default") -> str:
	return (
		"<td class='cover-kpi'>"
		f"<div class='cover-kpi-label'>{_esc(label)}</div>"
		f"<div class='cover-kpi-value accent-{accent}'>{_esc(value)}</div>"
		"</td>"
	)


def _kpi_cell(label: str, value: str, meta: str = "", accent: str = "default") -> str:
	return (
		"<td class='kpi-cell'>"
		f"<div class='kpi-label'>{_esc(label)}</div>"
		f"<div class='kpi-value accent-{accent}'>{_esc(value)}</div>"
		f"<div class='kpi-meta'>{_esc(meta)}</div>"
		"</td>"
	)


def _ladder_table(kpi: dict) -> str:
	rows = [
		("CM1 · Gross Margin", kpi["cm1"], kpi["cm1_pct"], "revenue − COGS"),
		("CM2 · Before marketing", kpi["cm2"], kpi["cm2_pct"], "− freight, packaging, txn, schemes"),
		("CM3 · After marketing", kpi["cm3"], kpi["cm3_pct"], "− marketing"),
		("EBITDA", kpi["ebitda"], kpi["ebitda_pct"], "− salary, rent, utilities, admin"),
		("EBIT", kpi["ebit"], kpi["ebit_pct"], "− D&A"),
		("PBT", kpi["pbt"], kpi["pbt_pct"], "− interest"),
		("PAT", kpi["pat"], kpi["pat_pct"], "− income tax"),
	]
	body = "".join(
		f"<tr><td class='tier'>{_esc(label)}</td>"
		f"<td class='num'>{_inr(abs_val)}</td>"
		f"<td class='num pct {'pos' if pct >= 0 else 'neg'}'>{_pct(pct)}</td>"
		f"<td class='hint'>{_esc(hint)}</td></tr>"
		for label, abs_val, pct, hint in rows
	)
	return (
		"<table class='ladder'>"
		"<thead><tr><th>Tier</th><th class='num'>₹</th><th class='num'>%</th><th>Definition</th></tr></thead>"
		f"<tbody>{body}</tbody></table>"
	)


def _pnl_table(kpi: dict, c2: dict, c3: dict, li: dict, rev: float) -> str:
	lines: list[tuple[str, float, bool, str]] = [
		("Gross Revenue", kpi["total_revenue"], True, ""),
		("  COGS (incoming_rate × qty)", -(kpi["total_revenue"] - kpi["gross_profit"]), False, "incl. primary packaging via BOM"),
		("CM1 · Gross Margin", kpi["cm1"], True, ""),
		("  Outbound Freight", -c2["freight_total"], False, f"{len(c2['freight_accounts'])} acc"),
		("  Packaging & Dispatch", -c2["packaging_total"], False, f"{len(c2['packaging_accounts'])} acc"),
		("  Transaction Fees & Commissions", -c2["commission_total"], False, f"{len(c2['commission_accounts'])} acc"),
		("  Trade Schemes & Rebates", -c2["scheme_total"], False, f"{len(c2['scheme_accounts'])} acc"),
		("CM2 · Before marketing", kpi["cm2"], True, ""),
		("  Marketing — online + offline", -c3["marketing_total"], False, f"{len(c3['marketing_accounts'])} acc"),
		("CM3 · After marketing", kpi["cm3"], True, ""),
		("  Salary & Payroll", -li["salary_total"], False, f"{len(li['salary_accounts'])} acc"),
		("  Rent & Lease", -li["rent_total"], False, f"{len(li['rent_accounts'])} acc"),
		("  Utilities & Communications", -li["utility_total"], False, f"{len(li['utility_accounts'])} acc"),
		("  Admin & General Overhead", -li["admin_total"], False, f"{len(li['admin_accounts'])} acc"),
		("EBITDA", kpi["ebitda"], True, ""),
		("  Depreciation & Amortization", -li["da_total"], False, f"{len(li['da_accounts'])} acc"),
		("EBIT", kpi["ebit"], True, ""),
		("  Interest & Finance Cost", -li["interest_total"], False, f"{len(li['interest_accounts'])} acc"),
		("PBT", kpi["pbt"], True, ""),
		("  Income Tax Expense", -li["tax_total"], False, f"{len(li['tax_accounts'])} acc"),
		("PAT", kpi["pat"], True, ""),
	]
	body = "".join(
		f"<tr class='{'tier' if is_tier else 'deduction'}'>"
		f"<td>{_esc(label)}</td>"
		f"<td class='num'>{_inr(amount)}</td>"
		f"<td class='num pct'>{_pct(amount / rev * 100)}</td>"
		f"<td class='hint'>{_esc(detail)}</td></tr>"
		for label, amount, is_tier, detail in lines
	)
	return (
		"<table class='pnl'>"
		"<thead><tr><th>Line</th><th class='num'>₹</th><th class='num'>% Rev</th><th>Detail</th></tr></thead>"
		f"<tbody>{body}</tbody></table>"
	)


def _monthly_table(rows: list[dict]) -> str:
	if not rows:
		return "<div class='empty'>No monthly data in period.</div>"
	body = "".join(
		f"<tr><td>{_esc(r['month'])} {r['year']}</td>"
		f"<td class='num'>{r['b2c']:.1f}</td>"
		f"<td class='num'>{r['b2b']:.1f}</td>"
		f"<td class='num'>{r['other']:.1f}</td>"
		f"<td class='num tier'>{r['total']:.1f}</td></tr>"
		for r in rows
	)
	return (
		"<table class='compact'>"
		"<thead><tr><th>Month</th><th class='num'>B2C</th><th class='num'>B2B</th><th class='num'>Other</th><th class='num'>Total</th></tr></thead>"
		f"<tbody>{body}</tbody></table>"
	)


def _category_table(rows: list[dict]) -> str:
	total = sum(r["revenue"] for r in rows) or 1
	body = "".join(
		f"<tr><td>{_esc(r['category'])}</td>"
		f"<td class='num'>{int(r['sku_count']):,}</td>"
		f"<td class='num'>{r['revenue']:.1f}</td>"
		f"<td class='num pct'>{_pct(r['revenue'] / total * 100)}</td></tr>"
		for r in rows
	)
	return (
		"<table class='compact'>"
		"<thead><tr><th>Category</th><th class='num'>SKUs</th><th class='num'>₹L</th><th class='num'>% Mix</th></tr></thead>"
		f"<tbody>{body}</tbody></table>"
	)


def _sku_table(rows: list[dict]) -> str:
	if not rows:
		return "<div class='empty'>No SKU data in period.</div>"
	body = "".join(
		f"<tr><td class='num muted'>{i + 1}</td>"
		f"<td>{_esc(r['item_code'])}<div class='hint truncate'>{_esc(r['item_name'])}</div></td>"
		f"<td>{_esc(r['item_group'])}</td>"
		f"<td class='num'>{int(r['qty']):,}</td>"
		f"<td class='num'>{_inr(r['revenue'])}</td>"
		f"<td class='num {'pos' if r['gm_pct'] > 5 else 'neg' if r['gm_pct'] < 0 else ''}'>{_pct(r['gm_pct'])}</td>"
		f"<td class='num {'pos' if r['cm3_pct'] > 2 else 'neg' if r['cm3_pct'] < 0 else ''}'>{_pct(r['cm3_pct'])}</td></tr>"
		for i, r in enumerate(rows)
	)
	return (
		"<table class='sku'>"
		"<thead><tr><th class='num'>#</th><th>SKU</th><th>Category</th><th class='num'>Units</th>"
		"<th class='num'>Revenue</th><th class='num'>GM %</th><th class='num'>CM3 %</th></tr></thead>"
		f"<tbody>{body}</tbody></table>"
	)


def _allocation_footer(alloc: dict) -> str:
	return (
		"<div class='footer-note'>"
		f"Allocations (% of revenue): freight {alloc['freight_pct']:.2f}%, "
		f"packaging {alloc['packaging_pct']:.2f}%, "
		f"txn fees {alloc['commission_pct']:.2f}%, "
		f"marketing {alloc['marketing_pct']:.2f}%. Basis: {_esc(alloc['basis'])}."
		"</div>"
	)


def _segment_summary_table(res: dict) -> str:
	rows = res.get("segments", [])
	totals = res.get("totals", {})
	total_rev = totals.get("revenue") or 1
	body_rows: list[str] = []
	for s in rows:
		share = s["revenue"] / total_rev * 100
		yoy = _pct(s["yoy_pct"]) if s["yoy_pct"] is not None else "—"
		body_rows.append(
			f"<tr><td>{_esc(s['segment'])}</td>"
			f"<td class='num'>{s['customers']:,}</td>"
			f"<td class='num'>{s['invoices']:,}</td>"
			f"<td class='num'>{_inr(s['revenue'])}</td>"
			f"<td class='num pct'>{_pct(share)}</td>"
			f"<td class='num pct {'pos' if (s['yoy_pct'] or 0) >= 0 else 'neg'}'>{_esc(yoy)}</td>"
			f"<td class='num {'pos' if s['gm_pct'] > 5 else 'neg' if s['gm_pct'] < 0 else ''}'>{_pct(s['gm_pct'])}</td></tr>"
		)
	# Totals row
	body_rows.append(
		f"<tr class='tier'><td>Total</td>"
		f"<td class='num'>{int(totals.get('customers', 0)):,}</td>"
		f"<td class='num'>{int(totals.get('invoices', 0)):,}</td>"
		f"<td class='num'>{_inr(totals.get('revenue', 0))}</td>"
		f"<td class='num pct'>100%</td><td class='num pct'>—</td>"
		f"<td class='num'>{_pct(totals.get('gm_pct', 0))}</td></tr>"
	)
	body = "".join(body_rows)
	return (
		"<table class='compact'>"
		"<thead><tr><th>Segment</th><th class='num'>Cust</th><th class='num'>Inv</th>"
		"<th class='num'>Revenue</th><th class='num'>% Mix</th><th class='num'>YoY</th>"
		"<th class='num'>GM %</th></tr></thead>"
		f"<tbody>{body}</tbody></table>"
	)


def _segment_top_customers(res: dict) -> str:
	rows = res.get("segments", [])[:6]  # cap to avoid pagination overflow
	blocks: list[str] = []
	for s in rows:
		if not s["top_customers"]:
			continue
		seg_rev = s["revenue"] or 1
		customer_rows = "".join(
			f"<tr><td>{_esc(c['customer_name'])}</td>"
			f"<td class='num muted'>{int(c['invoices']):,}</td>"
			f"<td class='num'>{_inr(c['revenue'])}</td>"
			f"<td class='num pct'>{_pct(c['revenue'] / seg_rev * 100)}</td></tr>"
			for c in s["top_customers"][:3]
		)
		blocks.append(
			f"<div class='mini-card'>"
			f"<div class='mini-head'>{_esc(s['segment'])}"
			f"<span class='mini-meta'>· {_inr(s['revenue'])}</span></div>"
			f"<table class='mini-tbl'>{customer_rows}</table>"
			f"</div>"
		)
	if not blocks:
		return ""
	return "<div class='sub-header'>Top 3 customers per segment</div>" + "<div class='mini-grid'>" + "".join(blocks) + "</div>"


def _yoy_chip(cur: float, prior: float) -> str:
	if not prior:
		return ""
	pct = (cur - prior) / prior * 100
	sign = "↑" if pct >= 0 else "↓"
	return f"{sign}{abs(pct):.0f}% YoY"


def _inr(value: float | int | None) -> str:
	if value is None:
		return "—"
	v = float(value)
	absv = abs(v)
	if absv >= 1_00_00_000:
		return f"₹{v / 1_00_00_000:.1f}Cr"
	if absv >= 1_00_000:
		return f"₹{round(v / 1_00_000):,}L"
	if absv >= 1_000:
		return f"₹{v / 1_000:.1f}K"
	return f"₹{round(v):,}"


def _pct(value: float | None, digits: int = 2) -> str:
	if value is None:
		return "—"
	return f"{value:.{digits}f}%"


def _esc(value) -> str:
	return _html.escape(str(value), quote=True) if value is not None else ""


_PDF_CSS = """
<style>
@page { size: A4; margin: 14mm 14mm 16mm 14mm; }
* { box-sizing: border-box; }
body {
	font-family: "Helvetica", "Arial", sans-serif;
	color: #1a1810;
	background: #ffffff;
	font-size: 10.5px;
	line-height: 1.45;
	margin: 0;
	padding: 0;
}
section { page-break-inside: avoid; }
.page { padding: 0; }
.page-break { page-break-before: always; }
.spacer-16 { height: 16px; }

/* Cover */
.cover {
	page-break-after: always;
	padding: 32mm 4mm 0 4mm;
	background: #f4f1e8;
	min-height: 240mm;
	border-left: 4px solid #1F5F33;
}
.cover-eyebrow {
	font-size: 10px;
	letter-spacing: 0.2em;
	text-transform: uppercase;
	color: #1F5F33;
	font-weight: 700;
	margin-bottom: 6mm;
}
.cover-title {
	font-family: "Times New Roman", serif;
	font-size: 32pt;
	line-height: 1.1;
	margin: 0 0 5mm 0;
	color: #1a1810;
}
.cover-period { font-size: 11.5px; color: #574f3e; font-weight: 600; }
.cover-tags { margin-top: 4mm; }
.tag {
	display: inline-block;
	font-size: 9.5px;
	font-weight: 600;
	padding: 3px 10px;
	border-radius: 10px;
	margin-right: 4px;
}
.tag-green { background: #dfeed9; color: #1F5F33; }
.tag-amber { background: #f8e7c9; color: #96560f; }
.tag-blue { background: #dae5f1; color: #2c4e80; }
.cover-spacer { height: 20mm; }
.cover-kpis {
	width: 100%;
	border-collapse: separate;
	border-spacing: 6px 0;
	margin-top: 12mm;
}
.cover-kpi {
	width: 16.6%;
	background: #ffffff;
	border: 1px solid rgba(40,30,15,0.12);
	padding: 5mm 4mm;
	border-radius: 6px;
	vertical-align: top;
}
.cover-kpi-label {
	font-size: 8.5px;
	letter-spacing: 0.1em;
	text-transform: uppercase;
	color: #7a6f5d;
	font-weight: 600;
	margin-bottom: 3mm;
}
.cover-kpi-value {
	font-size: 16pt;
	font-weight: 700;
	color: #1a1810;
	letter-spacing: -0.02em;
}
.cover-meta {
	/* Keep this in the cover's flow — ``position: absolute`` escaped wkhtmltopdf's
	 * page boundaries and repeated the timestamp on every subsequent page. */
	margin-top: 38mm;
	font-size: 9px;
	color: #9e9382;
}
.accent-green { color: #1F5F33; }
.accent-amber { color: #96560f; }
.accent-red { color: #7A1F1F; }
.accent-blue { color: #2c4e80; }

/* Section header */
.section-head {
	display: block;
	border-bottom: 1px solid rgba(40,30,15,0.15);
	padding-bottom: 3mm;
	margin-bottom: 4mm;
}
.section-title {
	font-family: "Times New Roman", serif;
	font-size: 18pt;
	font-weight: 700;
	color: #1a1810;
	line-height: 1.1;
}
.section-sub {
	font-size: 10px;
	color: #7a6f5d;
	margin-top: 1mm;
	font-weight: 600;
}
.sub-header {
	font-size: 10.5px;
	font-weight: 700;
	color: #1a1810;
	margin: 6mm 0 2mm 0;
	padding-bottom: 1mm;
	border-bottom: 1px solid rgba(40,30,15,0.08);
}

/* KPI grid */
.kpi-grid {
	width: 100%;
	border-collapse: separate;
	border-spacing: 4px;
}
.kpi-cell {
	width: 33.3%;
	background: #faf7f0;
	border: 1px solid rgba(40,30,15,0.12);
	padding: 3mm;
	vertical-align: top;
	border-radius: 4px;
}
.kpi-label {
	font-size: 8.5px;
	letter-spacing: 0.08em;
	text-transform: uppercase;
	color: #7a6f5d;
	font-weight: 600;
	margin-bottom: 1mm;
}
.kpi-value {
	font-size: 13pt;
	font-weight: 700;
	color: #1a1810;
	line-height: 1.1;
}
.kpi-meta {
	font-size: 9px;
	color: #574f3e;
	margin-top: 1mm;
}

/* Tables */
table.ladder, table.pnl, table.compact, table.sku, table.mini-tbl {
	width: 100%;
	border-collapse: collapse;
	font-size: 9.5px;
}
table.ladder th, table.pnl th, table.compact th, table.sku th {
	text-align: left;
	background: #1F5F33;
	color: #ffffff;
	font-weight: 600;
	font-size: 9px;
	letter-spacing: 0.06em;
	text-transform: uppercase;
	padding: 4px 6px;
}
table.ladder td, table.pnl td, table.compact td, table.sku td {
	padding: 3px 6px;
	border-bottom: 1px solid rgba(40,30,15,0.08);
	color: #1a1810;
}
table.ladder tr.tier td, table.pnl tr.tier td, .tier {
	background: #ebe6d9;
	font-weight: 700;
}
table.pnl tr.deduction td { color: #574f3e; }
.num { text-align: right; }
.pct { font-variant-numeric: tabular-nums; }
.pos { color: #1F5F33; }
.neg { color: #7A1F1F; }
.muted { color: #7a6f5d; }
.hint { color: #7a6f5d; font-size: 9px; }
.truncate {
	max-width: 58mm;
	overflow: hidden;
	text-overflow: ellipsis;
	white-space: nowrap;
}
.footer-note {
	margin-top: 3mm;
	font-size: 9px;
	color: #7a6f5d;
	font-style: italic;
}
.empty {
	padding: 8mm;
	text-align: center;
	color: #9e9382;
	font-size: 10px;
	font-style: italic;
}

/* Server-side SVG charts. Inline width 100% so wkhtmltopdf scales them to the
 * printable width while keeping the viewBox aspect ratio. Small top margin
 * separates chart from the section sub-header. */
svg.chart {
	display: block;
	margin: 2mm 0 4mm 0;
	max-width: 100%;
}

/* Mini cards for segments */
.mini-grid {
	display: block;
}
.mini-card {
	display: inline-block;
	width: 48%;
	margin: 1% 1% 3mm 0;
	border: 1px solid rgba(40,30,15,0.12);
	background: #ffffff;
	border-radius: 4px;
	padding: 3mm;
	vertical-align: top;
	page-break-inside: avoid;
}
.mini-head {
	font-weight: 700;
	font-size: 10px;
	color: #1a1810;
	margin-bottom: 2mm;
	padding-bottom: 1mm;
	border-bottom: 1px solid rgba(40,30,15,0.08);
}
.mini-meta {
	font-weight: 400;
	font-size: 9px;
	color: #7a6f5d;
	margin-left: 4px;
}
table.mini-tbl td { border-bottom: none; padding: 2px 4px; font-size: 9px; }
</style>
"""


# ── Server-side PNG charts (Pillow) ───────────────────────────────────────────
# wkhtmltopdf 0.12 silently drops inline SVG, so we rasterize to PNG server-side
# and embed as data URLs. Colours mirror the dashboard so the PDF stays visually
# consistent with the on-screen Recharts. Pillow ships with Frappe — no new dep.

_CH_GREEN = (31, 95, 51)
_CH_GREEN_LT = (31, 95, 51, 58)  # 22% alpha
_CH_AMBER = (150, 86, 15)
_CH_RED = (122, 31, 31)
_CH_BLUE = (44, 78, 128)
_CH_MUTED = (122, 111, 93)
_CH_GRID = (40, 30, 15, 20)
_CH_TEXT = (26, 24, 16)
_CH_WHITE = (255, 255, 255)
_CH_SURFACE = (250, 247, 240)

# 2× render resolution → downscale in the PDF via CSS width. Keeps charts crisp
# in print without blowing up file size.
_SCALE = 2


def _load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
	paths = [
		"/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
		if bold
		else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
		"/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf"
		if bold
		else "/usr/share/fonts/dejavu/DejaVuSans.ttf",
	]
	for p in paths:
		if os.path.exists(p):
			return ImageFont.truetype(p, size * _SCALE)
	return ImageFont.load_default()


def _png_data_url(img: Image.Image) -> str:
	buf = io.BytesIO()
	img.save(buf, format="PNG", optimize=True)
	return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("ascii")


def _img_tag(img: Image.Image, width_px: int) -> str:
	"""Render at 2× internally, display at 1× in the PDF so print output stays sharp."""
	return f'<img class="chart" src="{_png_data_url(img)}" style="width:{width_px}px;max-width:100%;">'


def _text_width(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> float:
	if hasattr(draw, "textbbox"):
		bbox = draw.textbbox((0, 0), text, font=font)
		return bbox[2] - bbox[0]
	return draw.textlength(text, font=font)




def _svg_monthly_chart(rows: list[dict], width: int = 720, height: int = 200) -> str:
	"""Monthly B2C area + B2B line + Total line, rendered to PNG for wkhtmltopdf."""
	if not rows:
		return "<div class='empty'>No monthly data.</div>"

	W, H = width * _SCALE, height * _SCALE
	img = Image.new("RGB", (W, H), _CH_WHITE)
	draw = ImageDraw.Draw(img, "RGBA")
	f_axis = _load_font(9)
	f_legend = _load_font(9)

	pad_l, pad_r, pad_t, pad_b = 44 * _SCALE, 12 * _SCALE, 8 * _SCALE, 32 * _SCALE
	chart_w = W - pad_l - pad_r
	chart_h = H - pad_t - pad_b
	n = len(rows)
	step = chart_w / max(n - 1, 1)
	max_val = max((r.get("total") or 0) for r in rows) or 1

	def xy(i: int, v: float) -> tuple[float, float]:
		x = pad_l + i * step
		y = pad_t + chart_h - (v / max_val) * chart_h
		return x, y

	# Grid + Y labels at 0 / 33% / 66% / 100%
	for frac, label in [(0, "0"), (0.33, f"{max_val * 0.33:.0f}"), (0.66, f"{max_val * 0.66:.0f}"), (1, f"{max_val:.0f}L")]:
		y = pad_t + chart_h - frac * chart_h
		draw.line([(pad_l, y), (pad_l + chart_w, y)], fill=_CH_GRID, width=1)
		draw.text((pad_l - 6 * _SCALE, y - 5 * _SCALE), label, font=f_axis, fill=_CH_MUTED, anchor="rm")

	# B2C area (fill + outline)
	b2c_pts = [xy(i, r.get("b2c") or 0) for i, r in enumerate(rows)]
	area_pts = b2c_pts + [(pad_l + chart_w, pad_t + chart_h), (pad_l, pad_t + chart_h)]
	draw.polygon(area_pts, fill=_CH_GREEN_LT)
	if len(b2c_pts) > 1:
		draw.line(b2c_pts, fill=_CH_GREEN, width=2 * _SCALE, joint="curve")

	# B2B line
	b2b_pts = [xy(i, r.get("b2b") or 0) for i, r in enumerate(rows)]
	if len(b2b_pts) > 1:
		draw.line(b2b_pts, fill=_CH_AMBER, width=2 * _SCALE, joint="curve")

	# Total dashed
	total_pts = [xy(i, r.get("total") or 0) for i, r in enumerate(rows)]
	for a, b in zip(total_pts, total_pts[1:]):
		# manual dashed — Pillow has no native dash pattern
		dx, dy = b[0] - a[0], b[1] - a[1]
		length = (dx * dx + dy * dy) ** 0.5 or 1
		ux, uy = dx / length, dy / length
		dash = 6 * _SCALE
		gap = 4 * _SCALE
		d = 0.0
		while d < length:
			x1 = a[0] + ux * d
			y1 = a[1] + uy * d
			x2 = a[0] + ux * min(d + dash, length)
			y2 = a[1] + uy * min(d + dash, length)
			draw.line([(x1, y1), (x2, y2)], fill=_CH_TEXT, width=1 * _SCALE)
			d += dash + gap

	# X-axis month labels
	tick = 1 if n <= 8 else 2
	for i, r in enumerate(rows):
		if i % tick != 0 and i != n - 1:
			continue
		x = pad_l + i * step
		draw.text((x, pad_t + chart_h + 6 * _SCALE), str(r.get("month", "")), font=f_axis, fill=_CH_MUTED, anchor="mt")

	# Legend
	lx = pad_l
	ly = H - 12 * _SCALE
	draw.rectangle([lx, ly, lx + 10 * _SCALE, ly + 3 * _SCALE], fill=_CH_GREEN)
	draw.text((lx + 14 * _SCALE, ly), "B2C", font=f_legend, fill=_CH_MUTED, anchor="lm")
	lx += 46 * _SCALE
	draw.rectangle([lx, ly, lx + 10 * _SCALE, ly + 3 * _SCALE], fill=_CH_AMBER)
	draw.text((lx + 14 * _SCALE, ly), "B2B", font=f_legend, fill=_CH_MUTED, anchor="lm")
	lx += 46 * _SCALE
	draw.rectangle([lx, ly, lx + 10 * _SCALE, ly + 1 * _SCALE], fill=_CH_TEXT)
	draw.text((lx + 14 * _SCALE, ly), "Total", font=f_legend, fill=_CH_MUTED, anchor="lm")

	return _img_tag(img, width)


def _svg_cm_ladder(kpi: dict, width: int = 720, height: int = 200) -> str:
	"""CM ladder bars: tier %-of-revenue, centred on zero."""
	tiers = [
		("CM1", kpi["cm1_pct"]),
		("CM2", kpi["cm2_pct"]),
		("CM3", kpi["cm3_pct"]),
		("EBITDA", kpi["ebitda_pct"]),
		("EBIT", kpi["ebit_pct"]),
		("PBT", kpi["pbt_pct"]),
		("PAT", kpi["pat_pct"]),
	]
	W, H = width * _SCALE, height * _SCALE
	img = Image.new("RGB", (W, H), _CH_WHITE)
	draw = ImageDraw.Draw(img)
	f_axis = _load_font(9)
	f_label = _load_font(9, bold=True)
	f_value = _load_font(9, bold=True)

	pad_l, pad_r, pad_t, pad_b = 36 * _SCALE, 12 * _SCALE, 16 * _SCALE, 30 * _SCALE
	chart_w = W - pad_l - pad_r
	chart_h = H - pad_t - pad_b
	y_zero = pad_t + chart_h / 2

	max_abs = max(abs(v) for _, v in tiers) or 1
	span = max_abs * 1.15

	def y_for(v: float) -> float:
		return y_zero - (v / span) * (chart_h / 2)

	# Gridlines + labels
	for v, label in [(span, f"{span:.1f}%"), (0, "0"), (-span, f"-{span:.1f}%")]:
		y = y_for(v)
		col = (60, 45, 25, 110) if v == 0 else _CH_GRID
		draw.line([(pad_l, y), (pad_l + chart_w, y)], fill=col, width=1)
		draw.text((pad_l - 6 * _SCALE, y - 5 * _SCALE), label, font=f_axis, fill=_CH_MUTED, anchor="rm")

	n = len(tiers)
	bar_w = chart_w / n * 0.58
	gap = chart_w / n - bar_w

	for i, (label, pct) in enumerate(tiers):
		bx = pad_l + i * (bar_w + gap) + gap / 2
		if pct >= 0:
			by = y_for(pct)
			bh = y_zero - by
		else:
			by = y_zero
			bh = y_for(pct) - y_zero
		fill = _CH_GREEN if pct >= 0 else _CH_RED
		draw.rectangle([bx, by, bx + bar_w, by + max(bh, 1.5 * _SCALE)], fill=fill)

		# Value label — above positive bars, below negative
		if pct >= 0:
			ty = by - 4 * _SCALE
			anchor = "mb"
		else:
			ty = by + bh + 4 * _SCALE
			anchor = "mt"
		draw.text((bx + bar_w / 2, ty), f"{pct:.2f}%", font=f_value, fill=_CH_TEXT, anchor=anchor)

		# X-axis tier label
		draw.text(
			(bx + bar_w / 2, pad_t + chart_h + 10 * _SCALE),
			label,
			font=f_label,
			fill=_CH_MUTED,
			anchor="mt",
		)

	return _img_tag(img, width)


def _svg_category_bars(rows: list[dict], width: int = 720) -> str:
	"""Horizontal bars for top categories, category name on the left."""
	if not rows:
		return ""
	top = rows[:10]
	max_val = max(r["revenue"] for r in top) or 1

	row_h = 18
	pad_l, pad_r, pad_t = 120, 80, 8
	height = len(top) * row_h + pad_t + 12
	W, H = width * _SCALE, height * _SCALE

	img = Image.new("RGB", (W, H), _CH_WHITE)
	draw = ImageDraw.Draw(img)
	f_label = _load_font(10)
	f_value = _load_font(9, bold=True)

	chart_w = W - (pad_l + pad_r) * _SCALE

	for i, r in enumerate(top):
		y = (pad_t + i * row_h + row_h / 2) * _SCALE
		bar_len = (r["revenue"] / max_val) * chart_w
		# category label (right-aligned)
		draw.text(((pad_l - 6) * _SCALE, y), _truncate(r["category"], 22), font=f_label, fill=_CH_TEXT, anchor="rm")
		# bar
		draw.rectangle(
			[pad_l * _SCALE, y - 7 * _SCALE, pad_l * _SCALE + max(bar_len, 2), y + 7 * _SCALE],
			fill=_CH_GREEN,
		)
		# value
		draw.text(
			(pad_l * _SCALE + bar_len + 4 * _SCALE, y),
			f"₹{r['revenue']:.1f}L",
			font=f_value,
			fill=_CH_TEXT,
			anchor="lm",
		)

	return _img_tag(img, width)


def _svg_segment_bars(res: dict, width: int = 720) -> str:
	"""Segment revenue bars, with % of total on the right. (unassigned) coloured amber."""
	rows = res.get("segments", [])
	if not rows:
		return ""
	top = rows[:10]
	total = res["totals"].get("revenue") or 1
	max_val = max(s["revenue"] for s in top) or 1

	row_h = 19
	pad_l, pad_r, pad_t = 140, 110, 8
	height = len(top) * row_h + pad_t + 12
	W, H = width * _SCALE, height * _SCALE

	img = Image.new("RGB", (W, H), _CH_WHITE)
	draw = ImageDraw.Draw(img)
	f_label = _load_font(10)
	f_value = _load_font(9, bold=True)
	f_share = _load_font(9)

	chart_w = W - (pad_l + pad_r) * _SCALE

	for i, s in enumerate(top):
		y = (pad_t + i * row_h + row_h / 2) * _SCALE
		bar_len = (s["revenue"] / max_val) * chart_w
		share = s["revenue"] / total * 100
		is_un = s["segment"] == "(unassigned)"
		fill = _CH_AMBER if is_un else _CH_BLUE
		draw.text(
			((pad_l - 6) * _SCALE, y),
			_truncate(s["segment"], 20),
			font=f_label,
			fill=_CH_TEXT,
			anchor="rm",
		)
		draw.rectangle(
			[pad_l * _SCALE, y - 7 * _SCALE, pad_l * _SCALE + max(bar_len, 2), y + 7 * _SCALE],
			fill=fill,
		)
		value_text = _inr(s["revenue"])
		draw.text(
			(pad_l * _SCALE + bar_len + 4 * _SCALE, y),
			value_text,
			font=f_value,
			fill=_CH_TEXT,
			anchor="lm",
		)
		# share in muted grey to the right of the value
		value_w = _text_width(draw, value_text, f_value)
		draw.text(
			(pad_l * _SCALE + bar_len + 4 * _SCALE + value_w + 6 * _SCALE, y),
			f"· {share:.1f}%",
			font=f_share,
			fill=_CH_MUTED,
			anchor="lm",
		)

	return _img_tag(img, width)


def _truncate(text: str, max_chars: int) -> str:
	return text if len(text) <= max_chars else text[: max_chars - 1] + "…"
